import re
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from schemas import SystemMetric, UploadedFileData

QM_FILENAME = re.compile(r"QM-.+\.txt")
WORKBOOK_FILENAME = re.compile(r"MRM_.+\.xlsx")
WORKBOOK_COLUMNS = {
    "monitoring_metric": ["Monitoring Metric", "Metric"],
    "test_objective": ["Test Objective"],
    "calculation_method": [
        "Calculation Method/Formula",
        "Calcution Method/Formula",
        "Calculation Method / Formula",
    ],
}
EMPTY_METRICS = {"any other(s)"}


def read_user_inputs(
    qm_file: UploadedFileData,
    workbook_file: UploadedFileData,
) -> tuple[str, list[SystemMetric]]:
    return _read_qm(qm_file), _read_workbook(workbook_file)


def _read_qm(upload: UploadedFileData) -> str:
    _validate_filename(upload.filename, QM_FILENAME, "QM-*.txt")
    if not upload.content:
        raise ValueError("The QM text file is empty.")
    try:
        content = upload.content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValueError("The QM text file must use UTF-8 encoding.") from None
    if not content.strip():
        raise ValueError("The QM text file contains no readable text.")
    return content


def _read_workbook(upload: UploadedFileData) -> list[SystemMetric]:
    _validate_filename(upload.filename, WORKBOOK_FILENAME, "MRM_*.xlsx")
    if not upload.content:
        raise ValueError("The developer workbook is empty.")
    try:
        workbook = load_workbook(BytesIO(upload.content), read_only=True, data_only=False)
    except Exception:
        raise ValueError("The developer workbook could not be read as XLSX.") from None

    try:
        rows = workbook.active.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            raise ValueError("The developer workbook is empty.")
        header_names = [str(value).strip() if value is not None else "" for value in headers]
        positions = {
            field: _column_position(header_names, aliases, field)
            for field, aliases in WORKBOOK_COLUMNS.items()
        }

        metrics = []
        names = set()
        for row in rows:
            name = _cell(row, positions["monitoring_metric"])
            objective = _cell(row, positions["test_objective"])
            calculation = _cell(row, positions["calculation_method"])
            normalized_name = name.casefold()

            if not name or (normalized_name in EMPTY_METRICS and not objective and not calculation):
                continue
            if normalized_name in names:
                raise ValueError(f"Developer workbook contains duplicate Metric '{name}'.")
            names.add(normalized_name)
            metrics.append(
                SystemMetric(
                    monitoring_metric=name,
                    test_objective=objective,
                    calculation_method=calculation,
                )
            )

        if not metrics:
            raise ValueError("The developer workbook contains no Metric rows.")
        return metrics
    finally:
        workbook.close()


def _validate_filename(filename: str, pattern: re.Pattern[str], expected: str) -> None:
    if Path(filename).name != filename or pattern.fullmatch(filename) is None:
        raise ValueError(f"Select a file named {expected}.")


def _column_position(headers: list[str], aliases: list[str], field: str) -> int:
    positions = [index for index, header in enumerate(headers) if header in aliases]
    if not positions:
        raise ValueError(f"Developer workbook is missing the {field} column.")
    if len(positions) > 1:
        raise ValueError(f"Developer workbook has more than one {field} column.")
    return positions[0]


def _cell(row: tuple[object, ...], position: int) -> str:
    if position >= len(row) or row[position] is None:
        return ""
    value = str(row[position]).strip()
    return "" if value == '""' else value
