import re
from collections.abc import Sequence
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from mrm_review.schemas import (
    DeveloperMetric,
    MetricCatalogItem,
    MetricCategory,
    ReviewPackage,
    UploadedFileData,
)

QM_FILENAME = re.compile(r"QM-.+\.txt")
WORKBOOK_FILENAME = re.compile(r"MRM_.+\.xlsx")
EMPTY_METRIC_PLACEHOLDERS = {"any other(s)"}

DEVELOPER_COLUMN_ALIASES = {
    "metric": ("Monitoring Metric", "Metric"),
    "test_objective": ("Test Objective",),
    "calculation_method": (
        "Calculation Method/Formula",
        "Calcution Method/Formula",
        "Calculation Method / Formula",
    ),
}


def read_review_package(
    qm_files: Sequence[UploadedFileData],
    workbook_files: Sequence[UploadedFileData],
    catalog_path: Path,
) -> ReviewPackage:
    qm_file = _only_file(qm_files, "QM text file")
    workbook_file = _only_file(workbook_files, "developer workbook")

    source_text = _read_qm_text(qm_file)
    developer_metrics = _read_developer_metrics(workbook_file)
    catalog = _read_metric_catalog(catalog_path)

    return ReviewPackage(
        source_text=source_text,
        catalog=catalog,
        developer_metrics=developer_metrics,
    )


def _only_file(files: Sequence[UploadedFileData], label: str) -> UploadedFileData:
    if len(files) != 1:
        raise ValueError(f"Select exactly one {label}.")
    return files[0]


def _read_qm_text(upload: UploadedFileData) -> str:
    _validate_filename(upload.filename, QM_FILENAME, "QM-*.txt")
    if not upload.content:
        raise ValueError("The QM text file is empty.")

    try:
        content = upload.content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise ValueError("The QM text file must use UTF-8 encoding.") from None

    if not content.strip():
        raise ValueError("The QM text file contains no readable text.")
    return f"SOURCE FILE: {upload.filename}\n{content}"


def _read_developer_metrics(upload: UploadedFileData) -> list[DeveloperMetric]:
    _validate_filename(upload.filename, WORKBOOK_FILENAME, "MRM_*.xlsx")
    if not upload.content:
        raise ValueError("The developer workbook is empty.")

    try:
        workbook = load_workbook(BytesIO(upload.content), read_only=True, data_only=False)
    except Exception:
        raise ValueError("The developer workbook could not be read as XLSX.") from None

    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        first_row = next(rows, None)
        if first_row is None:
            raise ValueError("The developer workbook is empty.")

        headers = [str(value).strip() if value is not None else "" for value in first_row]
        positions = {
            field: _find_column(headers, aliases, field)
            for field, aliases in DEVELOPER_COLUMN_ALIASES.items()
        }

        metrics: list[DeveloperMetric] = []
        metric_names: set[str] = set()
        for row in rows:
            name = _cell_text(row, positions["metric"])
            test_objective = _cell_text(row, positions["test_objective"])
            calculation_method = _cell_text(row, positions["calculation_method"])

            if not name:
                continue

            normalized_name = name.casefold()
            if (
                normalized_name in EMPTY_METRIC_PLACEHOLDERS
                and not test_objective
                and not calculation_method
            ):
                continue
            if normalized_name in metric_names:
                raise ValueError(f"Developer workbook contains duplicate Metric '{name}'.")
            metric_names.add(normalized_name)
            metrics.append(
                DeveloperMetric(
                    name=name,
                    test_objective=test_objective,
                    calculation_method=calculation_method,
                )
            )

        if not metrics:
            raise ValueError("The developer workbook contains no Metric rows.")
        return metrics
    finally:
        workbook.close()


def _validate_filename(filename: str, pattern: re.Pattern[str], expected: str) -> None:
    if Path(filename).name != filename or pattern.fullmatch(filename) is None:
        raise ValueError(f"Select a file named {expected}.")


def _find_column(headers: list[str], aliases: tuple[str, ...], field: str) -> int:
    positions = [index for index, header in enumerate(headers) if header in aliases]
    if not positions:
        raise ValueError(f"Developer workbook is missing the {field} column.")
    if len(positions) > 1:
        raise ValueError(f"Developer workbook has more than one {field} column.")
    return positions[0]


def _cell_text(row: tuple[object, ...], position: int) -> str:
    if position >= len(row) or row[position] is None:
        return ""
    value = str(row[position]).strip()
    return "" if value == '""' else value


def _read_metric_catalog(path: Path) -> list[MetricCategory]:
    if not path.is_file():
        raise FileNotFoundError("metrics/metrics.md was not found.")

    categories: list[MetricCategory] = []
    category_name = ""
    section = ""
    subcategories: list[str] = []
    applications: list[str] = []
    metric_names: list[str] = []

    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()

        if line.startswith("## "):
            if category_name:
                _append_category(
                    categories,
                    _build_metric_category(
                        category_name,
                        subcategories,
                        applications,
                        metric_names,
                    ),
                )
            category_name = _category_name(line[3:])
            section = ""
            subcategories = []
            applications = []
            metric_names = []
            continue

        if line.startswith(("* **", "- **")):
            section = _plain_markdown_text(line[2:])
            continue

        if not category_name or not line.startswith(("* ", "- ")):
            continue

        value = _plain_markdown_text(line[2:])
        if section == "Main Subcategories":
            _append_catalog_value(subcategories, value, category_name, section)
        elif section == "Applications":
            _append_catalog_value(applications, value, category_name, section)
        elif section == "Metrics":
            _append_catalog_value(metric_names, value, category_name, section)

    if category_name:
        _append_category(
            categories,
            _build_metric_category(
                category_name,
                subcategories,
                applications,
                metric_names,
            ),
        )

    if not categories:
        raise ValueError("metrics.md must contain at least one category.")
    return categories


def _build_metric_category(
    name: str,
    subcategories: list[str],
    applications: list[str],
    metric_names: list[str],
) -> MetricCategory:
    missing_sections = []
    if not subcategories:
        missing_sections.append("Main Subcategories")
    if not applications:
        missing_sections.append("Applications")
    if not metric_names:
        missing_sections.append("Metrics")
    if missing_sections:
        raise ValueError(f"Catalog category '{name}' is missing: {', '.join(missing_sections)}.")

    return MetricCategory(
        name=name,
        subcategories=subcategories,
        applications=applications,
        metrics=[
            MetricCatalogItem(name=metric, guidance=f"Approved under {name}")
            for metric in metric_names
        ],
    )


def _append_category(categories: list[MetricCategory], category: MetricCategory) -> None:
    if category.name.casefold() in {item.name.casefold() for item in categories}:
        raise ValueError(f"Catalog contains duplicate category '{category.name}'.")
    categories.append(category)


def _append_catalog_value(
    values: list[str],
    value: str,
    category: str,
    section: str,
) -> None:
    if not value:
        return
    if value.casefold() in {item.casefold() for item in values}:
        raise ValueError(f"Catalog category '{category}' has duplicate {section} value '{value}'.")
    values.append(value)


def _category_name(value: str) -> str:
    name = _plain_markdown_text(value)
    prefix, separator, remainder = name.partition(". ")
    if separator and prefix.isdigit():
        return remainder
    return name


def _plain_markdown_text(value: str) -> str:
    return value.replace("**", "").strip()
